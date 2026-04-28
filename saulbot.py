from flask import Flask, request
import requests
import os
import json
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()

app = Flask(__name__)

VERIFY_TOKEN = os.getenv("VERIFY_TOKEN")
ACCESS_TOKEN = os.getenv("WHATSAPP_ACCESS_TOKEN")
PHONE_NUMBER_ID = os.getenv("PHONE_NUMBER_ID")
CONFIG_FILE = "config.json"

ESTADOS_USUARIO = {}

CAMPOS_DIRECCION = [
    ("calle", "calle"),
    ("numero", "numero"),
    ("colonia", "colonia"),
    ("municipio", "municipio"),
    ("estado", "estado"),
    ("codigo_postal", "codigo postal"),
    ("pais", "pais"),
]

CONFIG_DATA = {}


def cargar_config():
    if not os.path.exists(CONFIG_FILE):
        return {"clientes": {}}
    with open(CONFIG_FILE, encoding="utf-8") as f:
        return json.load(f)


def obtener_config_cliente(numero):
    return CONFIG_DATA.get("clientes", {}).get(numero, {})


def guardar_cliente_config(numero, direccion):
    clientes = CONFIG_DATA.setdefault("clientes", {})
    cliente = clientes.setdefault(numero, {})
    cliente["direccion"] = direccion
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(CONFIG_DATA, f, ensure_ascii=False, indent=2)


CONFIG_DATA = cargar_config()

PALABRAS = CONFIG_DATA.get("palabras_clave", {})
MENSAJES = CONFIG_DATA.get("mensajes", {})
NEGOCIO = CONFIG_DATA.get("negocio", {})
PRODUCTOS = CONFIG_DATA.get("productos", {})

# =========================
# FUNCION PARA ENVIAR MENSAJES
# =========================
def normalizar_numero(numero: str) -> str:
    numero = "".join(ch for ch in str(numero) if ch.isdigit())

    # Caso comun para Mexico: 521XXXXXXXXXX -> 52XXXXXXXXXX
    if numero.startswith("521") and len(numero) == 13:
        numero = "52" + numero[3:]

    return numero


def enviar_mensaje(numero, mensaje):
    numero = normalizar_numero(numero)
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }

    data = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "text",
        "text": {"body": mensaje}
    }

    response = requests.post(url, headers=headers, json=data)
    print(f"[WhatsApp API] To: {numero} | Status: {response.status_code} | Body: {response.text}")

def generar_menu():
    menu = "Nuestros productos:\n\n"
    for numero, producto in PRODUCTOS.items():
        menu += f"{numero}. {producto['nombre']} - ${producto['precio']}\n"
    menu += "\nResponde con el número del producto que deseas ordenar."
    return menu

def valor_normalizado(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def extraer_componentes_direccion(lat, lon):
    url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=jsonv2&addressdetails=1"
    headers = {
        "User-Agent": "BotWhatsapp/1.0"
    }
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        data = response.json()
        address = data.get("address", {})

        componentes = {
            "calle": valor_normalizado(address.get("road") or address.get("pedestrian") or address.get("footway") or address.get("residential")),
            "numero": valor_normalizado(address.get("house_number")),
            "colonia": valor_normalizado(address.get("suburb") or address.get("neighbourhood") or address.get("quarter") or address.get("city_district") or address.get("hamlet")),
            "municipio": valor_normalizado(address.get("city") or address.get("town") or address.get("municipality") or address.get("county") or address.get("village")),
            "estado": valor_normalizado(address.get("state")),
            "codigo_postal": valor_normalizado(address.get("postcode")),
            "pais": valor_normalizado(address.get("country")),
        }

        return componentes
    except Exception as e:
        print(f"Error al obtener direccion: {e}")
        return {campo: "" for campo, _ in CAMPOS_DIRECCION}


def formatear_direccion(componentes):
    return ", ".join(
        valor_normalizado(componentes.get(campo))
        for campo, _ in CAMPOS_DIRECCION
        if valor_normalizado(componentes.get(campo))
    )


def obtener_campos_faltantes(componentes):
    return [campo for campo, _ in CAMPOS_DIRECCION if not valor_normalizado(componentes.get(campo))]


def obtener_nombre_campo(campo):
    for clave, nombre in CAMPOS_DIRECCION:
        if clave == campo:
            return nombre
    return campo


def solicitar_siguiente_campo(numero, estado):
    faltantes = obtener_campos_faltantes(estado["direccion"])

    if not faltantes:
        direccion_completa = formatear_direccion(estado["direccion"])
        estado["paso"] = "confirmando_direccion"
        enviar_mensaje(numero, f"Esta es la direccion completa:\n{direccion_completa}")
        enviar_mensaje(numero, "Confirmas que es correcta? (si/no)")
        return

    siguiente = faltantes[0]
    estado["paso"] = "completando_direccion"
    estado["campo_pendiente"] = siguiente
    enviar_mensaje(numero, f"Me falta este dato de tu direccion: {obtener_nombre_campo(siguiente)}. Por favor escribelo.")


def manejar_texto_segun_estado(numero, texto, estado):
    paso = estado.get("paso")
    
    if paso == "seleccionando_producto":
        producto = PRODUCTOS.get(texto.strip())
        if producto:
            estado["pedido"] = f"{producto['nombre']} - ${producto['precio']}"
            estado["paso"] = None
            enviar_mensaje(numero, f"Anotado: {estado['pedido']}")
            enviar_mensaje(numero, MENSAJES.get("pedido_recibido", ""))
        else:
            enviar_mensaje(numero, "Por favor responde con el número del producto.")
        return True

    elif paso == "completando_direccion":
        campo = estado.get("campo_pendiente")
        if campo:
            estado["direccion"][campo] = texto.strip()
        solicitar_siguiente_campo(numero, estado)
        return True

    elif paso == "confirmando_direccion":
        respuesta = texto.strip().lower()

        if respuesta in {"si", "sí", "correcta", "confirmo"}:
            direccion_completa = formatear_direccion(estado["direccion"])
            guardar_pedido(numero, estado.get("pedido", "Pedido pendiente"), direccion_completa)
            guardar_cliente_config(numero, estado["direccion"])
            enviar_mensaje(numero, f"Perfecto. Direccion confirmada:\n{direccion_completa}")
            enviar_mensaje(numero, MENSAJES.get("pedido_confirmado", ""))
            estado.clear()
            return True

        elif respuesta in {"no", "incorrecta"}:
            estado["direccion"] = {campo: "" for campo, _ in CAMPOS_DIRECCION}
            estado["paso"] = "completando_direccion"
            estado["campo_pendiente"] = "calle"
            enviar_mensaje(numero, "De acuerdo. Vamos a capturar tu direccion manualmente.")
            enviar_mensaje(numero, "Escribe tu calle.")
            return True

        enviar_mensaje(numero, "Responde con si o no para confirmar la direccion.")
        return True

    return False

# =========================
# OBTENER DIRECCION DESDE COORDENADAS
# =========================
def obtener_direccion(lat, lon):
    componentes = extraer_componentes_direccion(lat, lon)
    direccion = formatear_direccion(componentes)
    return direccion or "Direccion no encontrada", componentes

# =========================
# GUARDAR PEDIDO EN EXCEL
# =========================
def guardar_pedido(numero, pedido, direccion):
    archivo = NEGOCIO.get("archivo_pedidos", "pedidos.xlsx")

    if not os.path.exists(archivo):
        wb = Workbook()
        ws = wb.active
        ws.append(["Numero", "Pedido", "Direccion"])
        wb.save(archivo)

    wb = load_workbook(archivo)
    ws = wb.active

    ws.append([numero, pedido, direccion])
    wb.save(archivo)

# =========================
# WEBHOOK
# =========================
@app.route("/webhook", methods=["GET", "POST"])
def webhook():
    if request.method == "GET":
        token = request.args.get("hub.verify_token")
        challenge = request.args.get("hub.challenge")

        if token == VERIFY_TOKEN:
            return challenge
        else:
            return "Error", 403

    data = request.get_json(silent=True) or {}

    try:
        entry = data.get("entry", [])
        changes = entry[0].get("changes", []) if entry else []
        value = changes[0].get("value", {}) if changes else {}
        mensajes = value.get("messages", [])

        if not mensajes:
            return "evento sin mensaje", 200

        mensaje = mensajes[0]
        numero = mensaje.get("from")
        estado_usuario = ESTADOS_USUARIO.setdefault(numero, {})

        if not numero:
            return "remitente no encontrado", 200

        cliente_config = obtener_config_cliente(numero)
        if cliente_config:
            estado_usuario.setdefault("direccion", cliente_config.get("direccion", {}))
            estado_usuario.setdefault("pedido", "Pedido pendiente")

        # TEXTO
        if mensaje.get("type") == "text":
            texto_original = mensaje.get("text", {}).get("body", "")
            texto = texto_original.lower()

            if manejar_texto_segun_estado(numero, texto_original, estado_usuario):
                return "ok", 200

            if any(p in texto for p in PALABRAS.get("saludo", [])):
                enviar_mensaje(numero, MENSAJES.get("bienvenida", ""))

            elif any(p in texto for p in PALABRAS.get("pedido", [])):
                estado_usuario["paso"] = "seleccionando_producto"
                enviar_mensaje(numero, generar_menu())

            elif any(p in texto for p in PALABRAS.get("domicilio", [])):
                direccion_guardada = estado_usuario.get("direccion")
                if direccion_guardada and all(valor_normalizado(direccion_guardada.get(c)) for c, _ in CAMPOS_DIRECCION):
                    estado_usuario["paso"] = "confirmando_direccion"
                    direccion_completa = formatear_direccion(direccion_guardada)
                    enviar_mensaje(numero, f"Tengo esta direccion guardada para ti:\n{direccion_completa}")
                    enviar_mensaje(numero, "¿Deseas usarla? (si/no)")
                else:
                    enviar_mensaje(numero, MENSAJES.get("pedir_ubicacion", ""))

            elif any(p in texto for p in PALABRAS.get("recoger", [])):
                enviar_mensaje(numero, MENSAJES.get("recoger_confirmado", ""))

            else:
                enviar_mensaje(numero, "No entendi tu mensaje")         

        # UBICACION
        elif mensaje.get("type") == "location":
            lat = mensaje.get("location", {}).get("latitude")
            lon = mensaje.get("location", {}).get("longitude")

            if lat is None or lon is None:
                return "ubicacion incompleta", 200

            direccion, componentes = obtener_direccion(lat, lon)
            estado_usuario["direccion"] = componentes
            estado_usuario["paso"] = "direccion_recibida"
            estado_usuario.setdefault("pedido", "Pedido pendiente")

            if direccion != "Direccion no encontrada":
                enviar_mensaje(numero, f"Detecte esta direccion desde tu ubicacion:\n{direccion}")

            solicitar_siguiente_campo(numero, estado_usuario)

    except Exception as e:
        print("Error:", e)

    return "ok", 200

# =========================
# MAIN
# =========================
if __name__ == "__main__":
    app.run(port=5000, debug=True)